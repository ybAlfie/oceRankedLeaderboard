#!/usr/bin/env python3
"""
OCE Slapshot: Rebound Leaderboard Builder

Reads a spreadsheet of player names + Slapshot IDs, fetches each player's
current rank from slapshot.gg's public JSON endpoints, and optionally
crawls match history to discover OCE players missing from the list.

Usage:
    python oce_leaderboard.py [xlsx_path] [output_dir]
    python oce_leaderboard.py --no-discover            # Phase 1 only (known IDs)
    python oce_leaderboard.py --max-discovery 100      # Cap BFS growth

Outputs (in output_dir/):
    - oce_leaderboard.csv        Sorted leaderboard of all matched players
    - discovered_not_on_list.csv OCE players found in match history not in your spreadsheet
    - unmatched_names.csv        Names in your spreadsheet we couldn't find an ID for
    - raw_data.json              Full API responses (useful for debugging / later enrichment)
"""

import argparse
import csv
import json
import sys
import time
from collections import deque
from pathlib import Path

import openpyxl
import requests


# ---- Config ----

BASE_URL = "https://slapshot.gg"
OCE_REGION = "oce-east"
DELAY_SECONDS = 1           # Pause between requests; RateTracker handles the ~10/30s cap
DEFAULT_MAX_DISCOVERY = 500 # Cap on new players to discover via BFS

SEASON3_START = "2025-09-23"          # Only count matches from this date onward
CASUAL_MATCH_TYPE = "casual"          # Ranked game mode; excludes pond, customs
MIN_OCE_SEASON3_MATCHES = 2          # Min qualifying matches to appear in not_on_list

HEADERS = {
    # Present as a normal browser; referer matches what Firefox actually sends
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://slapshot.gg/",
}


# ---- HTTP with rate limit awareness ----

class RateTracker:
    def __init__(self):
        self.remaining = None
        self.reset_at = None

    def update(self, resp):
        if "x-ratelimit-remaining" in resp.headers:
            try:
                self.remaining = int(resp.headers["x-ratelimit-remaining"])
                self.reset_at = float(resp.headers["x-ratelimit-reset"])
            except (ValueError, TypeError):
                pass

    def wait_if_needed(self):
        if self.remaining is not None and self.remaining <= 2 and self.reset_at:
            wait = max(0, self.reset_at - time.time()) + 1
            if wait > 0:
                print(f"   [rate limit low, sleeping {wait:.1f}s]")
                time.sleep(wait)


def fetch_json(url, rate, session):
    rate.wait_if_needed()
    time.sleep(DELAY_SECONDS)

    try:
        resp = session.get(url, headers=HEADERS, timeout=30)
    except requests.RequestException as e:
        print(f"   ! network error on {url}: {e}")
        return None

    rate.update(resp)

    if resp.status_code == 429:
        print("   ! 429 Too Many Requests, backing off 60s")
        time.sleep(60)
        return fetch_json(url, rate, session)

    if resp.status_code in (404, 403):
        return None

    if resp.status_code >= 400:
        print(f"   ! HTTP {resp.status_code} on {url}")
        return None

    try:
        return resp.json()
    except ValueError:
        print(f"   ! non-JSON response from {url}")
        return None


def get_player(pid, rate, session):
    return fetch_json(f"{BASE_URL}/api/game/players/{pid}", rate, session)


def get_ranked(pid, rate, session):
    return fetch_json(f"{BASE_URL}/api/game/players/{pid}/ranked", rate, session)


# ---- Input / helpers ----

def load_players(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    players = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        pid = str(row[1]).strip() if len(row) > 1 and row[1] is not None else None
        if pid == "None":
            pid = None
        players.append({"name": name, "id": pid})
    return players


def extract_match_player_ids(match):
    """Pull all player IDs out of a match entry. Defensive about field names."""
    ids = set()
    game_stats = (match or {}).get("game_stats") or {}
    for p in game_stats.get("players", []) or []:
        pid = (
            p.get("game_user_id")
            or p.get("player_id")
            or p.get("id")
            or p.get("user_id")
        )
        if pid is not None:
            ids.add(str(pid))
    return ids


def is_qualifying_match(match):
    """Return True if a match counts as an OCE casual Season 3 game."""
    return (
        match.get("region") == OCE_REGION
        and match.get("match_type") == CASUAL_MATCH_TYPE
        and (match.get("created") or "") >= SEASON3_START
    )


def count_qualifying_matches(player_data):
    """Count OCE Season 3 matches (any mode) to gauge if a player is OCE-based."""
    return sum(1 for m in (player_data or {}).get("match_history", []) or []
               if m.get("region") == OCE_REGION
               and (m.get("created") or "") >= SEASON3_START)


# ---- HTML output ----

def build_html(players):
    data_json = json.dumps([
        {
            "pos": i + 1,
            "name": p["name"],
            "rank": p["rank"] or "",
            "rating": p["rating"] if p["rating"] is not None else "",
            "matches": p["matches_played"] if p["matches_played"] is not None else 0,
            "highest_rank": p["highest_rank"] or "",
            "highest_rating": p["highest_rating"] if p["highest_rating"] is not None else "",
            "id": p["id"],
        }
        for i, p in enumerate(players)
    ], ensure_ascii=False)

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>OCE Slapshot Leaderboard</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ background: #0f1117; color: #e2e8f0; font-family: system-ui, sans-serif; padding: 24px; }}
  h1 {{ font-size: 1.5rem; font-weight: 700; margin-bottom: 4px; color: #fff; }}
  .subtitle {{ color: #64748b; font-size: 0.85rem; margin-bottom: 20px; }}
  .controls {{ display: flex; gap: 10px; flex-wrap: wrap; margin-bottom: 18px; align-items: center; }}
  .search {{ background: #1e2230; border: 1px solid #2d3348; color: #e2e8f0; padding: 7px 12px;
             border-radius: 6px; font-size: 0.9rem; width: 220px; }}
  .search::placeholder {{ color: #475569; }}
  .search:focus {{ outline: none; border-color: #4f8ef7; }}
  .toggle {{ background: #1e2230; border: 1px solid #2d3348; color: #94a3b8; padding: 7px 14px;
             border-radius: 6px; font-size: 0.85rem; cursor: pointer; transition: all .15s; }}
  .toggle.active {{ background: #4f8ef7; border-color: #4f8ef7; color: #fff; }}
  .count {{ color: #64748b; font-size: 0.85rem; margin-left: auto; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 0.88rem; }}
  thead th {{ background: #161b27; color: #94a3b8; font-weight: 600; text-align: left;
              padding: 10px 14px; border-bottom: 1px solid #2d3348; white-space: nowrap;
              cursor: pointer; user-select: none; position: sticky; top: 0; z-index: 1; }}
  thead th:hover {{ color: #e2e8f0; }}
  thead th.sorted {{ color: #4f8ef7; }}
  thead th .arrow {{ margin-left: 4px; font-size: 0.7rem; opacity: 0.6; }}
  thead th.sorted .arrow {{ opacity: 1; }}
  tbody tr {{ border-bottom: 1px solid #1a1f2e; transition: background .1s; }}
  tbody tr:hover {{ background: #1a2035; }}
  td {{ padding: 9px 14px; }}
  td.pos {{ color: #475569; width: 48px; }}
  td.name {{ font-weight: 500; color: #fff; }}
  td.num {{ text-align: right; font-variant-numeric: tabular-nums; }}
  .rank-badge {{ display: inline-block; padding: 2px 8px; border-radius: 4px;
                 font-size: 0.78rem; font-weight: 600; }}
  .rank-Platinum\\ I  {{ background:#2a1f6e; color:#a78bfa; }}
  .rank-Gold\\ III    {{ background:#2d2200; color:#f59e0b; }}
  .rank-Gold\\ II     {{ background:#2d2200; color:#fbbf24; }}
  .rank-Gold\\ I      {{ background:#2d2200; color:#fcd34d; }}
  .rank-Silver\\ III  {{ background:#1e2530; color:#94a3b8; }}
  .rank-Silver\\ II   {{ background:#1e2530; color:#cbd5e1; }}
  .rank-Silver\\ I    {{ background:#1e2530; color:#e2e8f0; }}
  .rank-Bronze\\ III  {{ background:#2a1a0e; color:#d97706; }}
  .rank-Bronze\\ II   {{ background:#2a1a0e; color:#b45309; }}
  .rank-Bronze\\ I    {{ background:#2a1a0e; color:#92400e; }}
  .rank-Unranked      {{ background:#1a1f2e; color:#475569; }}
  .rank-             {{ background:#1a1f2e; color:#475569; }}
</style>
</head>
<body>
<h1>OCE Slapshot Leaderboard</h1>
<p class="subtitle">Season 3 &mdash; oce-east</p>
<div class="controls">
  <input class="search" type="text" placeholder="Search name..." id="search">
  <button class="toggle" id="toggleZero">Hide 0 matches</button>
  <button class="toggle" id="toggleUnranked">Hide unranked</button>
  <span class="count" id="count"></span>
</div>
<table>
  <thead>
    <tr>
      <th data-col="pos">#<span class="arrow">&#8597;</span></th>
      <th data-col="name">Name<span class="arrow">&#8597;</span></th>
      <th data-col="rank">Rank<span class="arrow">&#8597;</span></th>
      <th data-col="rating" class="sorted">Rating<span class="arrow">&#8595;</span></th>
      <th data-col="matches">Matches<span class="arrow">&#8597;</span></th>
      <th data-col="highest_rank">Highest Rank<span class="arrow">&#8597;</span></th>
      <th data-col="highest_rating">Highest Rating<span class="arrow">&#8597;</span></th>
    </tr>
  </thead>
  <tbody id="tbody"></tbody>
</table>
<script>
const RANK_ORDER = [
  'Platinum I','Gold III','Gold II','Gold I',
  'Silver III','Silver II','Silver I',
  'Bronze III','Bronze II','Bronze I','Unranked',''
];
const rankVal = r => {{ const i = RANK_ORDER.indexOf(r); return i === -1 ? 99 : i; }};

const ALL = {data_json};

let sortCol = 'rating', sortAsc = false;
let hideZero = false, hideUnranked = false, searchTerm = '';

function filtered() {{
  return ALL.filter(p => {{
    if (hideZero && p.matches === 0) return false;
    if (hideUnranked && (p.rank === 'Unranked' || p.rank === '')) return false;
    if (searchTerm && !p.name.toLowerCase().includes(searchTerm)) return false;
    return true;
  }});
}}

function sorted(rows) {{
  return [...rows].sort((a, b) => {{
    let av = a[sortCol], bv = b[sortCol];
    if (sortCol === 'rank' || sortCol === 'highest_rank') {{
      av = rankVal(av); bv = rankVal(bv);
    }} else if (typeof av === 'string') {{
      av = av.toLowerCase(); bv = bv.toLowerCase();
    }}
    if (av === '' || av === null) av = sortAsc ? Infinity : -Infinity;
    if (bv === '' || bv === null) bv = sortAsc ? Infinity : -Infinity;
    return sortAsc ? (av > bv ? 1 : av < bv ? -1 : 0) : (av < bv ? 1 : av > bv ? -1 : 0);
  }});
}}

function render() {{
  const rows = sorted(filtered());
  document.getElementById('count').textContent = rows.length + ' players';
  const tbody = document.getElementById('tbody');
  tbody.innerHTML = rows.map((p, i) => `
    <tr>
      <td class="pos">${{i + 1}}</td>
      <td class="name">${{p.name}}</td>
      <td><span class="rank-badge rank-${{p.rank}}">${{p.rank || 'Unranked'}}</span></td>
      <td class="num">${{p.rating}}</td>
      <td class="num">${{p.matches}}</td>
      <td><span class="rank-badge rank-${{p.highest_rank}}">${{p.highest_rank || '—'}}</span></td>
      <td class="num">${{p.highest_rating}}</td>
    </tr>`).join('');
}}

document.querySelectorAll('thead th').forEach(th => {{
  th.addEventListener('click', () => {{
    const col = th.dataset.col;
    if (sortCol === col) sortAsc = !sortAsc;
    else {{ sortCol = col; sortAsc = col === 'name'; }}
    document.querySelectorAll('thead th').forEach(t => {{
      t.classList.remove('sorted');
      t.querySelector('.arrow').innerHTML = '&#8597;';
    }});
    th.classList.add('sorted');
    th.querySelector('.arrow').innerHTML = sortAsc ? '&#8593;' : '&#8595;';
    render();
  }});
}});

document.getElementById('toggleZero').addEventListener('click', function() {{
  hideZero = !hideZero; this.classList.toggle('active', hideZero); render();
}});
document.getElementById('toggleUnranked').addEventListener('click', function() {{
  hideUnranked = !hideUnranked; this.classList.toggle('active', hideUnranked); render();
}});
document.getElementById('search').addEventListener('input', function() {{
  searchTerm = this.value.toLowerCase(); render();
}});

render();
</script>
</body>
</html>"""


# ---- Main pipeline ----

def fetch_all_for_id(pid, rate, session, cache):
    if pid in cache:
        return cache[pid]
    player = get_player(pid, rate, session)
    ranked = get_ranked(pid, rate, session)
    cache[pid] = {"player": player, "ranked": ranked}
    return cache[pid]


def run(xlsx_path, output_dir, discover=True, max_discovery=DEFAULT_MAX_DISCOVERY, resume=False):
    players = load_players(xlsx_path)
    known_ids = [p["id"] for p in players if p["id"]]
    print(f"Loaded {len(players)} players; {len(known_ids)} have IDs, "
          f"{len(players) - len(known_ids)} don't\n")

    rate = RateTracker()
    session = requests.Session()

    # Load previous run's cache if resuming
    cache = {}
    raw_path = Path(output_dir) / "raw_data.json"
    if resume and raw_path.exists():
        with open(raw_path) as f:
            cache = json.load(f)
        print(f"Resumed cache: {len(cache)} IDs already fetched\n")

    # ---- Phase 1: known IDs ----
    print("=== Phase 1: fetching known IDs ===")
    for i, pid in enumerate(known_ids, 1):
        if pid in cache:
            print(f"[{i}/{len(known_ids)}] id={pid} (cached)")
            continue
        print(f"[{i}/{len(known_ids)}] id={pid}")
        fetch_all_for_id(pid, rate, session, cache)

    # ---- Phase 2: BFS discovery through match history ----
    discovered = set()
    if discover:
        print(f"\n=== Phase 2: BFS discovery (cap: {max_discovery}) ===")
        # When resuming, treat everything already in cache as visited/queued
        # so BFS explores from the frontier of the previous run
        visited = set(cache.keys()) if resume else set(known_ids)
        queue = deque(cache.keys()) if resume else deque(known_ids)

        while queue and len(discovered) < max_discovery:
            pid = queue.popleft()
            player_data = cache.get(pid, {}).get("player")
            if not player_data:
                continue

            for match in player_data.get("match_history", []) or []:
                if not is_qualifying_match(match):
                    continue
                for new_id in extract_match_player_ids(match):
                    if new_id in visited or new_id.startswith("bot-"):
                        continue
                    visited.add(new_id)
                    discovered.add(new_id)
                    print(f"[discover {len(discovered)}] id={new_id}")
                    fetch_all_for_id(new_id, rate, session, cache)
                    queue.append(new_id)
                    if len(discovered) >= max_discovery:
                        break
                if len(discovered) >= max_discovery:
                    break

    # ---- Auto-match discovered usernames to unknown names in the list ----
    print("\n=== Matching discovered usernames to unknown names ===")
    username_to_id = {}
    for pid, data in cache.items():
        pd = data.get("player") or {}
        uname = (pd.get("username") or "").strip().upper()
        if uname:
            # If multiple IDs share a name (rare), keep the one with most matches_played
            existing = username_to_id.get(uname)
            if existing is None:
                username_to_id[uname] = pid
            else:
                new_mp = ((data.get("ranked") or {}).get("matches_played") or 0)
                old_mp = ((cache[existing].get("ranked") or {}).get("matches_played") or 0)
                if new_mp > old_mp:
                    username_to_id[uname] = pid

    matched = 0
    for p in players:
        if not p["id"]:
            up = p["name"].upper()
            if up in username_to_id:
                p["id"] = username_to_id[up]
                matched += 1
                print(f"   + {p['name']} -> {p['id']}")
    print(f"Auto-matched {matched} of "
          f"{sum(1 for p in players if not p['id']) + matched} previously-unknown names")

    # Write matched IDs back into the xlsx so future runs treat them as known
    if matched > 0:
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        name_to_id = {p["name"]: p["id"] for p in players if p["id"]}
        for row in ws.iter_rows(min_row=2):
            if not row[0].value:
                continue
            name = str(row[0].value).strip()
            if name in name_to_id and (not row[1].value):
                row[1].value = name_to_id[name]
        wb.save(xlsx_path)
        print(f"Saved {matched} matched IDs back to {xlsx_path}")

    # ---- Build outputs ----
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    leaderboard = []
    for p in players:
        if not p["id"] or p["id"] not in cache:
            continue
        r = cache[p["id"]].get("ranked") or {}
        pd = cache[p["id"]].get("player") or {}
        leaderboard.append({
            "name": p["name"],
            "username": pd.get("username", ""),
            "id": p["id"],
            "rating": r.get("rating"),
            "rank": (r.get("rank") or {}).get("name", ""),
            "rank_key": (r.get("rank") or {}).get("key", ""),
            "highest_rating": r.get("highest_rating"),
            "highest_rank": (r.get("highest_rank") or {}).get("name", ""),
            "matches_played": r.get("matches_played"),
            "season": r.get("season", ""),
        })
    # Sort: ranked players first (by rating desc), then unranked
    leaderboard.sort(key=lambda x: (x["rating"] is None, -(x["rating"] or 0)))

    with open(output_dir / "oce_leaderboard.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=[
            "position", "name", "username", "rank", "rating",
            "highest_rank", "highest_rating", "matches_played", "season", "id",
        ])
        w.writeheader()
        for i, p in enumerate(leaderboard, 1):
            w.writerow({"position": i, **{k: v for k, v in p.items() if k != "rank_key"}})

    # Discovered players NOT on the spreadsheet (candidates to add)
    names_on_list = {p["name"].upper() for p in players}
    not_on_list = []
    for pid, data in cache.items():
        pd = data.get("player") or {}
        uname = (pd.get("username") or "").strip()
        if not uname or uname.upper() in names_on_list:
            continue
        if count_qualifying_matches(pd) < MIN_OCE_SEASON3_MATCHES:
            continue
        r = data.get("ranked") or {}
        not_on_list.append({
            "username": uname,
            "id": pid,
            "rating": r.get("rating"),
            "rank": (r.get("rank") or {}).get("name", ""),
            "matches_played": r.get("matches_played"),
        })
    not_on_list.sort(key=lambda x: (x["rating"] is None, -(x["rating"] or 0)))

    with open(output_dir / "discovered_not_on_list.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=[
            "username", "id", "rank", "rating", "matches_played"
        ])
        w.writeheader()
        w.writerows(not_on_list)

    # ---- Full combined leaderboard (spreadsheet players + discovered) ----
    seen_ids = set()
    full_leaderboard = []
    for p in leaderboard:
        seen_ids.add(p["id"])
        full_leaderboard.append({
            "name": p["name"],
            "id": p["id"],
            "rank": p["rank"],
            "rating": p["rating"],
            "matches_played": p["matches_played"],
            "highest_rank": p["highest_rank"],
            "highest_rating": p["highest_rating"],
        })
    for p in not_on_list:
        if p["id"] in seen_ids:
            continue
        r = cache.get(p["id"], {}).get("ranked") or {}
        full_leaderboard.append({
            "name": p["username"],
            "id": p["id"],
            "rank": p["rank"],
            "rating": p["rating"],
            "matches_played": p["matches_played"],
            "highest_rank": (r.get("highest_rank") or {}).get("name", ""),
            "highest_rating": r.get("highest_rating"),
        })
    full_leaderboard.sort(key=lambda x: (x["rating"] is None, -(x["rating"] or 0)))

    with open(output_dir / "full_leaderboard.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=[
            "position", "name", "rank", "rating", "matches_played",
            "highest_rank", "highest_rating", "id",
        ])
        w.writeheader()
        for i, p in enumerate(full_leaderboard, 1):
            w.writerow({"position": i, **p})

    with open(output_dir / "leaderboard.html", "w", encoding="utf-8") as f:
        f.write(build_html(full_leaderboard))

    # Names in the spreadsheet we still couldn't match
    unmatched = [p["name"] for p in players if not p["id"]]
    with open(output_dir / "unmatched_names.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["name"])
        for n in unmatched:
            w.writerow([n])

    # Raw data saved first so a later CSV crash doesn't lose the cache
    with open(output_dir / "raw_data.json", "w") as f:
        json.dump(cache, f, indent=2)

    print(f"\nDone! Written to {output_dir}/")
    print(f"  leaderboard.html            open in browser")
    print(f"  full_leaderboard.csv        {len(full_leaderboard)} total OCE players")
    print(f"  oce_leaderboard.csv         {len(leaderboard)} spreadsheet players")
    print(f"  discovered_not_on_list.csv  {len(not_on_list)} OCE players not on your list")
    print(f"  unmatched_names.csv         {len(unmatched)} names still missing an ID")


# ---- CLI ----

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("xlsx", nargs="?", default="Leaderboard_Names_and_IDs.xlsx",
                        help="Path to the name/ID spreadsheet")
    parser.add_argument("output", nargs="?", default="output",
                        help="Directory to write CSVs and raw data to")
    parser.add_argument("--no-discover", action="store_true",
                        help="Skip Phase 2 (BFS discovery)")
    parser.add_argument("--max-discovery", type=int, default=DEFAULT_MAX_DISCOVERY,
                        help=f"Cap on discovered players (default: {DEFAULT_MAX_DISCOVERY})")
    parser.add_argument("--delay", type=float, default=DELAY_SECONDS,
                        help=f"Seconds between requests (default: {DELAY_SECONDS})")
    parser.add_argument("--resume", action="store_true",
                        help="Load output/raw_data.json and skip already-fetched IDs")
    args = parser.parse_args()

    DELAY_SECONDS = args.delay
    run(args.xlsx, args.output,
        discover=not args.no_discover,
        max_discovery=args.max_discovery,
        resume=args.resume)
